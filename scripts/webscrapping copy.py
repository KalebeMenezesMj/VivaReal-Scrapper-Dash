import os
import re
import time
import json
import random
import pandas as pd
from tqdm import tqdm
from bs4 import BeautifulSoup
from colorama import init, Fore, Style
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from supabase import create_client, Client
from dotenv import load_dotenv

# ====== CONFIGURAÇÃO ======
init(autoreset=True)
load_dotenv()

URL_LISTAGEM = "https://www.vivareal.com.br/venda/"

SUPABASE_URL = os.getenv('VITE_SUPABASE_URL')
SUPABASE_KEY = os.getenv('VITE_SUPABASE_ANON_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print(Fore.RED + "ERRO: Variáveis de ambiente do Supabase não encontradas!")
    exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ====== FUNÇÕES AUXILIARES ======
def limpar_console():
    os.system("cls" if os.name == "nt" else "clear")

def human_sleep(a=0.6, b=1.2):
    time.sleep(random.uniform(a, b))

def normalizar_preco(preco_str):
    if not preco_str or preco_str == "0":
        return 0
    preco_str = preco_str.replace(".", "").replace(",", "").strip()
    try:
        return int(preco_str)
    except:
        return 0

def extrair_valores(texto):
    dados = {}
    preco_match = re.search(r"R\$\s*([\d\.\,]+)", texto)
    dados['valor'] = preco_match.group(1).strip() if preco_match else "0"
    metragem_match = re.search(r"([\d\.]+)\s*m²", texto, re.IGNORECASE)
    dados['area_privativa'] = metragem_match.group(1) if metragem_match else "0"
    quartos_match = re.search(r"(\d+)\s*quartos?", texto, re.IGNORECASE)
    dados['dormitorio'] = quartos_match.group(1) if quartos_match else "0"
    banheiros_match = re.search(r"(\d+)\s*banheiros?", texto, re.IGNORECASE)
    dados['banheiro'] = banheiros_match.group(1) if banheiros_match else "0"
    vagas_match = re.search(r"(\d+)\s*vagas?", texto, re.IGNORECASE)
    dados['vaga'] = vagas_match.group(1) if vagas_match else "0"
    suites_match = re.search(r"(\d+)\s*suítes?", texto, re.IGNORECASE)
    dados['suite'] = suites_match.group(1) if suites_match else "0"
    andar_match = re.search(r"(\d+)(?:º)?\s*andar", texto, re.IGNORECASE)
    dados['andar'] = andar_match.group(1) if andar_match else "0"
    dados['piscina'] = "1" if re.search(r"piscinas?", texto, re.IGNORECASE) else "0"
    dados['varanda'] = "1" if re.search(r"varandas?", texto, re.IGNORECASE) else "0"
    dados['elevador'] = "1" if re.search(r"elevador", texto, re.IGNORECASE) else "0"
    return dados

def dividir_endereco(endereco_texto):
    if not endereco_texto or endereco_texto == "0":
        return "0", "0", "0", "0"
    m = re.match(r"^(.*?)\s*-\s*(.*?),\s*(.*?)\s*-\s*(.*)$", endereco_texto)
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).strip(), m.group(4).strip()
    return endereco_texto, "0", "0", "0"

def get_address_from_soup(soup):
    tag = soup.select_one('p[data-testid="location-address"], div[data-testid="location-address"]')
    if tag:
        return tag.get_text(" ", strip=True)
    tag = soup.select_one('span[itemprop="streetAddress"]')
    if tag:
        return tag.get_text(" ", strip=True)
    return "0"

def estilizar_excel(nome_arquivo):
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2
    ws.freeze_panes = "A2"
    wb.save(nome_arquivo)

# ====== COLETAR LINKS ======
def coletar_links_listagem(max_scrolls=5):
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(URL_LISTAGEM)
    time.sleep(3)
    for i in range(max_scrolls):
        driver.execute_script("window.scrollBy(0, window.innerHeight);")
        human_sleep(1.5, 2.5)
    anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/imovel/']")
    links = sorted(set(a.get_attribute("href") for a in anchors if a.get_attribute("href")))
    driver.quit()
    return links

# ====== EXTRAIR INFORMAÇÕES DE UMA PÁGINA ======
def extrair_informacoes(url):
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.maximize_window()
    driver.get(url)
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)
    html_completo = driver.page_source
    driver.quit()
    soup = BeautifulSoup(html_completo, "html.parser")
    texto = soup.get_text(" ", strip=True)
    dados = extrair_valores(texto)
    dados['valor'] = normalizar_preco(dados.pop('valor'))
    endereco = get_address_from_soup(soup)
    dados['endereco_completo'] = endereco
    rua, bairro, cidade, estado = dividir_endereco(endereco)
    dados['rua'], dados['bairro'], dados['cidade'], dados['uf'] = rua, bairro, cidade, estado
    if 'suítes' in dados:
        dados['suite'] = dados.pop('suíte')
    return dados

# ====== MAIN ======
if __name__ == "__main__":
    limpar_console()
    print(Fore.GREEN + Style.BRIGHT + "=== COLETOR DE DADOS VIVAREAL (Supabase) ===\n")

    max_scrolls = 6

    # Criar job
    try:
        job_response = supabase.table('scraping_jobs').insert({
            'status': 'running',
            'max_scrolls': max_scrolls,
            'started_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).execute()
        job_id = job_response.data[0]['id']
        print(Fore.GREEN + f"Job criado com ID: {job_id}\n")
    except Exception as e:
        print(Fore.RED + f"Erro ao criar job: {e}")
        exit(1)

    # Coletar links
    try:
        links = coletar_links_listagem(max_scrolls=max_scrolls)
        print(Fore.CYAN + f"Foram coletados {len(links)} links.\n")
        supabase.table('scraping_jobs').update({'links_found': len(links)}).eq('id', job_id).execute()
    except Exception as e:
        print(Fore.RED + f"Erro ao coletar links: {e}")
        supabase.table('scraping_jobs').update({
            'status': 'failed',
            'error_message': str(e),
            'completed_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).eq('id', job_id).execute()
        exit(1)

    resultados = []
    properties_scraped = 0

    for url in tqdm(links, desc=f"{Fore.CYAN}Processando páginas", unit="página"):
        try:
            dados = extrair_informacoes(url)
            dados['link'] = url
            dados['job_id'] = job_id
            resultados.append(dados)

            if len(resultados) >= 10:
                try:
                    supabase.table('properties').insert(resultados).execute()
                    properties_scraped += len(resultados)
                    supabase.table('scraping_jobs').update({
                        'properties_scraped': properties_scraped
                    }).eq('id', job_id).execute()
                    resultados = []
                except Exception as e:
                    print(Fore.RED + f"Erro ao inserir no banco: {e}")
        except Exception as e:
            print(Fore.RED + f"Erro ao processar {url}: {e}")

    if resultados:
        try:
            supabase.table('properties').insert(resultados).execute()
            properties_scraped += len(resultados)
        except Exception as e:
            print(Fore.RED + f"Erro ao inserir últimos registros: {e}")

    # Gerar Excel
    try:
        all_properties = supabase.table('properties').select('*').eq('job_id', job_id).execute()
        if all_properties.data:
            df = pd.DataFrame(all_properties.data)
            colunas = ['valor', 'area_privativa', 'dormitorio', 'banheiro', 'vaga', 'suite',
                       'andar', 'piscina', 'varanda', 'elevador',
                       'rua', 'bairro', 'cidade', 'uf', 'endereco_completo', 'link']
            df = df.reindex(columns=colunas, fill_value='0')
            output_file = f"resultados_{job_id}.xlsx"
            df.to_excel(output_file, index=False)
            estilizar_excel(output_file)
            print(Fore.GREEN + f"Backup salvo em {output_file}")
    except Exception as e:
        print(Fore.RED + f"Erro ao gerar Excel: {e}")

    # Finalizar job
    try:
        supabase.table('scraping_jobs').update({
            'status': 'completed',
            'properties_scraped': properties_scraped,
            'completed_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).eq('id', job_id).execute()
        print(Fore.GREEN + Style.BRIGHT + f"\n✓ Scraping concluído!")
        print(Fore.CYAN + f"Total de imóveis coletados: {properties_scraped}")
    except Exception as e:
        print(Fore.RED + f"Erro ao finalizar job: {e}")
