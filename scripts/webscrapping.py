import os
import re
import time
import json
import random
import pandas as pd
from tqdm import tqdm
from bs4 import BeautifulSoup
from colorama import init, Fore, Style
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from supabase import create_client, Client
from dotenv import load_dotenv
from urllib.parse import urlparse, urlunparse
import concurrent.futures


# ====== CONFIGURAÇÃO ======
init(autoreset=True)
load_dotenv()

URL_LISTAGEM = "https://www.vivareal.com.br/venda/sp/santos/bairros/santa-maria/apartamento_residencial/?transacao=venda&onde=%2CS%C3%A3o+Paulo%2CSantos%2C%2CSanta+Maria%2C%2C%2Cneighborhood%2CBR%3ESao+Paulo%3ENULL%3ESantos%3EBarrios%3ESanta+Maria%2C-23.940526%2C-46.370098%2C%3B%2CS%C3%A3o+Paulo%2CSantos%2C%2CAreia+Branca%2C%2C%2Cneighborhood%2CBR%3ESao+Paulo%3ENULL%3ESantos%3EBarrios%3EAreia+Branca%2C-23.946714%2C-46.373514%2C&tipos=apartamento_residencial&areaMaxima=132&areaMinima=33"

SUPABASE_URL = os.getenv('VITE_SUPABASE_URL')
SUPABASE_KEY = os.getenv('VITE_SUPABASE_ANON_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print(Fore.RED + "ERRO: Variáveis de ambiente do Supabase não encontradas!")
    exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ====== FUNÇÕES AUXILIARES ======
def limpar_console():
    os.system("cls" if os.name == "nt" else "clear")

def human_sleep(a=0.6, b=1.2):
    time.sleep(random.uniform(a, b))

def normalize_url(url: str) -> str:
    """Remove query parameters and fragments to create a canonical URL."""
    if not url:
        return ""
    parsed = urlparse(url)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, '', '', ''))

def extrair_tipo_imovel(url: str) -> str:
    PREFIXOS = [
        # Commercial Types
        'consultorio', 'galpao-deposito-armazem', 'imovel-comercial',
        'ponto-comercial', 'sala-comercial', 'predio-comercial',
        # Residential Types
        'edificio-residencial', 'casa-de-condominio', 'fazenda---sitio',
        'lote-terreno', 'apartamento', 'cobertura', 'sobrado',
        'kitnet', 'flat', 'casa'
    ]
    try:
        path = urlparse(url).path
        segmento = path.split('/imovel/')[1]
        for prefixo in PREFIXOS:
            if segmento.startswith(prefixo):
                return 'casa isolada' if prefixo == 'casa' else prefixo.replace('-', ' ')
        return 'nao informado'
    except (IndexError, AttributeError):
        return 'nao informado'

def to_int(value_str, default=0):
    if not value_str: return default
    try:
        return int(re.sub(r'[^\d]', '', value_str))
    except (ValueError, TypeError):
        return default

def to_float(value_str, default=0.0):
    if not value_str: return default
    try:
        # Handles formats like "1.500,50" or "1500"
        cleaned_str = re.sub(r'[^\d,]', '', value_str).replace(',', '.')
        return float(cleaned_str)
    except (ValueError, TypeError):
        return default

def to_bool(value_str):
    return value_str == "1"

def extrair_valores(texto):
    dados = {}
    preco_match = re.search(r"R\$\s*([\d\.\,]+)", texto)
    dados['valor'] = preco_match.group(1).strip() if preco_match else "0"
    metragem_match = re.search(r"([\d\.,]+)\s*m²", texto, re.IGNORECASE)
    dados['area_privativa'] = metragem_match.group(1) if metragem_match else "0"
    quartos_match = re.search(r"(\d+)\s*quartos?", texto, re.IGNORECASE)
    dados['dormitorio'] = quartos_match.group(1) if quartos_match else "0"
    banheiros_match = re.search(r"(\d+)\s*banheiros?", texto, re.IGNORECASE)
    dados['banheiro'] = banheiros_match.group(1) if banheiros_match else "0"
    vagas_match = re.search(r"(\d+)\s*vagas?", texto, re.IGNORECASE)
    dados['vaga'] = vagas_match.group(1) if vagas_match else "0"
    suites_match = re.search(r"(\d+)\s*suítes?", texto, re.IGNORECASE)
    dados['suite'] = suites_match.group(1) if suites_match else "0"
    andar_match = re.search(r"(\d+)(?:º)?\s*andar", texto, re.IGNORECASE)
    dados['andar'] = andar_match.group(1) if andar_match else "0"
    dados['piscina'] = "1" if re.search(r"piscinas?", texto, re.IGNORECASE) else "0"
    dados['varanda'] = "1" if re.search(r"varandas?", texto, re.IGNORECASE) else "0"
    dados['elevador'] = "1" if re.search(r"elevador", texto, re.IGNORECASE) else "0"
    return dados

def dividir_endereco(endereco_texto):
    if not endereco_texto or endereco_texto == "0":
        return "0", "0", "0", "0"
    # Matches "Street, Number - Neighborhood, City - UF"
    m = re.match(r"^(.*?)\s*-\s*(.*?),\s*(.*?)\s*-\s*(.{2})$", endereco_texto)
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).strip(), m.group(4).strip()
    return endereco_texto, "0", "0", "0"

def get_address_from_soup(soup):
    for selector in ['p[data-testid="location-address"]', 'div[data-testid="location-address"]', 'span[itemprop="streetAddress"]']:
        tag = soup.select_one(selector)
        if tag:
            return tag.get_text(" ", strip=True)
    return "0"

def estilizar_excel(nome_arquivo):
    try:
        wb = load_workbook(nome_arquivo)
        ws = wb.active
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        ws.freeze_panes = "A2"
        wb.save(nome_arquivo)
    except Exception as e:
        print(Fore.RED + f"Erro ao estilizar Excel: {e}")

# ====== COLETAR LINKS ======
def coletar_links_listagem():
    print(Fore.YELLOW + "Iniciando navegador para coleta de links em todas as páginas...")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    #options.add_argument("--headless")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    all_hrefs = set()
    page_count = 1

    try:
        print(Fore.CYAN + f"Acessando URL inicial: {URL_LISTAGEM}")
        driver.get(URL_LISTAGEM)
        human_sleep(3, 4)

        while True:
            print(f"\n{Fore.CYAN}--- Processando Página {page_count} ---")
            
            # Scroll to ensure all lazy-loaded elements are present
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            human_sleep(1.5, 2.5)

            anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/imovel/']")
            page_hrefs = {a.get_attribute("href") for a in anchors if a.get_attribute("href")}
            
            new_links_count = len(page_hrefs - all_hrefs)
            print(f"{Fore.WHITE}Encontrados {len(page_hrefs)} links nesta página ({new_links_count} novos).")
            all_hrefs.update(page_hrefs)

            try:
                # ATUALIZADO: Novo seletor para o link <a> da próxima página
                next_button = driver.find_element(By.CSS_SELECTOR, "a[aria-label='próxima página']")
                
                # ATUALIZADO: Nova verificação para o estado desabilitado
                if next_button.get_attribute('aria-disabled') == 'true':
                    print(Fore.GREEN + "Link 'Próxima página' desabilitado. Fim da navegação.")
                    break
                
                # Use JavaScript click to avoid interception issues
                driver.execute_script("arguments[0].click();", next_button)
                print(f"{Fore.YELLOW}Navegando para a página {page_count + 1}...")
                page_count += 1
                human_sleep(3, 5) # Wait for new page content

            except NoSuchElementException:
                print(Fore.GREEN + "Link 'Próxima página' não encontrado. Fim da navegação.")
                break
            except Exception as e:
                print(Fore.RED + f"Erro inesperado ao tentar paginar: {e}")
                break

        print(f"\n{Fore.GREEN+Style.BRIGHT}Coleta de links finalizada. Total de links únicos encontrados: {len(all_hrefs)}")
        return list(all_hrefs)
    finally:
        driver.quit()

# ====== EXTRAIR INFORMAÇÕES DE UMA PÁGINA (THREAD-SAFE) ======
def extrair_informacoes(url):
    driver = None
    try:
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.maximize_window()
        driver.get(url)
        time.sleep(2)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        
        soup = BeautifulSoup(driver.page_source, "html.parser")
        texto = soup.get_text(" ", strip=True)
        
        dados_brutos = extrair_valores(texto)
        
        dados_convertidos = {
            'valor': to_float(dados_brutos.get('valor')),
            'area_privativa': to_float(dados_brutos.get('area_privativa')),
            'dormitorio': to_int(dados_brutos.get('dormitorio')),
            'banheiro': to_int(dados_brutos.get('banheiro')),
            'vaga': to_int(dados_brutos.get('vaga')),
            'suite': to_int(dados_brutos.get('suite')),
            'andar': dados_brutos.get('andar', '0'), # Keep as text
            'piscina': to_bool(dados_brutos.get('piscina')),
            'varanda': to_bool(dados_brutos.get('varanda')),
            'elevador': to_bool(dados_brutos.get('elevador')),
            'tipo': extrair_tipo_imovel(url)
        }
        
        endereco = get_address_from_soup(soup)
        dados_convertidos['endereco_completo'] = endereco
        rua, bairro, cidade, uf = dividir_endereco(endereco)
        dados_convertidos.update({'rua': rua, 'bairro': bairro, 'cidade': cidade, 'uf': uf})
        
        return dados_convertidos
    except Exception as e:
        print(Fore.RED + f"\nErro ao processar {url}: {e}")
        return None
    finally:
        if driver:
            driver.quit()

# ====== MAIN ======
if __name__ == "__main__":
    limpar_console()
    print(Fore.GREEN + Style.BRIGHT + "=== COLETOR DE DADOS VIVAREAL (Paginado e Paralelo) ===\n")

    # Pre-cache the Chrome Driver to avoid race conditions in threads
    print(Fore.YELLOW + "Verificando e instalando o ChromeDriver, se necessário...")
    try:
        ChromeDriverManager().install()
        print(Fore.GREEN + "ChromeDriver está pronto.\n")
    except Exception as e:
        print(Fore.RED + f"Não foi possível instalar o ChromeDriver: {e}")
        exit(1)

    job_id = None

    try:
        job_response = supabase.table('scraping_jobs').insert({
            'status': 'running', 'started_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).execute()
        job_id = job_response.data[0]['id']
        print(Fore.GREEN + f"Job de scraping criado com ID: {job_id}\n")
    except Exception as e:
        print(Fore.RED + f"Erro ao criar job no Supabase: {e}"); exit(1)

    try:
        links_brutos = coletar_links_listagem()
        
        print(Fore.MAGENTA + "\n" + "="*50)
        print(Fore.MAGENTA + "NORMALIZANDO E FILTRANDO DUPLICADOS")
        print(Fore.MAGENTA + "="*50)
        
        links_unicos_scrape = {normalize_url(link) for link in links_brutos}
        print(f"{Fore.WHITE}Links únicos (normalizados) nesta varredura: {len(links_unicos_scrape)}")

        print(Fore.YELLOW + "Buscando links existentes no banco de dados...")
        existing_links_response = supabase.table('properties').select('link').execute()
        
        existing_links_db = {item['link'] for item in existing_links_response.data} if existing_links_response.data else set()
        print(f"{Fore.WHITE}Links existentes no banco: {len(existing_links_db)}")

        new_links_to_process = list(links_unicos_scrape - existing_links_db)
        print(f"{Fore.GREEN+Style.BRIGHT}Total de links NOVOS para processar: {len(new_links_to_process)}\n")

        supabase.table('scraping_jobs').update({
            'links_found': len(links_brutos), 'new_links_to_process': len(new_links_to_process)
        }).eq('id', job_id).execute()
    except Exception as e:
        print(Fore.RED + f"Erro ao coletar e filtrar links: {e}")
        supabase.table('scraping_jobs').update({
            'status': 'failed', 'error_message': str(e), 'completed_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).eq('id', job_id).execute(); exit(1)

    all_results = []
    if not new_links_to_process:
        print(Fore.GREEN + "Nenhum imóvel novo para adicionar.")
    else:
        print(Fore.CYAN + "Iniciando scraping paralelo com 5 workers...")
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_url = {executor.submit(extrair_informacoes, url): url for url in new_links_to_process}
            for future in tqdm(concurrent.futures.as_completed(future_to_url), total=len(new_links_to_process), desc=f"{Fore.CYAN}Processando imóveis", unit="imóvel"):
                dados = future.result()
                if dados:
                    dados['link'] = future_to_url[future]
                    dados['job_id'] = job_id
                    all_results.append(dados)

    properties_scraped = 0
    if all_results:
        print(f"\n{Fore.YELLOW}Scraping concluído. {len(all_results)} imóveis coletados. Salvando no banco...")
        BATCH_SIZE = 50
        for i in tqdm(range(0, len(all_results), BATCH_SIZE), desc=f"{Fore.CYAN}Salvando no Supabase", unit="lote"):
            batch = all_results[i:i + BATCH_SIZE]
            try:
                supabase.table('properties').upsert(
                    batch, 
                    on_conflict='link'
                ).execute()
                properties_scraped += len(batch)
                supabase.table('scraping_jobs').update({'properties_scraped': properties_scraped}).eq('id', job_id).execute()
            except Exception as e:
                print(Fore.RED + f"\nErro ao inserir/atualizar lote no banco: {e}")

    if all_results:
        try:
            df = pd.DataFrame(all_results)
            colunas = ['tipo', 'valor', 'area_privativa', 'dormitorio', 'banheiro', 'vaga', 'suite',
                       'andar', 'piscina', 'varanda', 'elevador',
                       'rua', 'bairro', 'cidade', 'uf', 'endereco_completo', 'link']
            df = df.reindex(columns=colunas, fill_value=0)
            output_file = f"resultados_{job_id}.xlsx"
            df.to_excel(output_file, index=False)
            estilizar_excel(output_file)
            print(Fore.GREEN + f"\nBackup dos novos imóveis salvo em {output_file}")
        except Exception as e:
            print(Fore.RED + f"Erro ao gerar Excel: {e}")

    try:
        supabase.table('scraping_jobs').update({
            'status': 'completed', 'properties_scraped': len(all_results), 'completed_at': time.strftime('%Y-%m-%d %H:%M:%S')
        }).eq('id', job_id).execute()
        print(Fore.GREEN + Style.BRIGHT + f"\n✓ Scraping concluído!")
        print(Fore.CYAN + f"Total de imóveis novos processados: {len(all_results)}")
    except Exception as e:
        print(Fore.RED + f"Erro ao finalizar job: {e}")
