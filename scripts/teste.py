import re
import json
import time
import os
import pandas as pd
from tqdm import tqdm
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from colorama import init, Fore, Style
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# inicializa colorama
init(autoreset=True)


def limpar_console():
    os.system("cls" if os.name == "nt" else "clear")


def extrair_valores(texto):
    dados = {}

    preco_match = re.search(r"R\$\s*([\d\.\,]+)", texto)
    dados['preco'] = preco_match.group(1).strip() if preco_match else "0"

    metragem_match = re.search(r"([\d\.]+)\s*m²", texto, re.IGNORECASE)
    dados['metragem'] = metragem_match.group(1) if metragem_match else "0"

    quartos_match = re.search(r"(\d+)\s*quartos?", texto, re.IGNORECASE)
    dados['quartos'] = quartos_match.group(1) if quartos_match else "0"

    banheiros_match = re.search(r"(\d+)\s*banheiros?", texto, re.IGNORECASE)
    dados['banheiros'] = banheiros_match.group(1) if banheiros_match else "0"

    vagas_match = re.search(r"(\d+)\s*vagas?", texto, re.IGNORECASE)
    dados['vagas'] = vagas_match.group(1) if vagas_match else "0"

    suites_match = re.search(r"(\d+)\s*suítes?", texto, re.IGNORECASE)
    dados['suítes'] = suites_match.group(1) if suites_match else "0"

    andar_match = re.search(r"(\d+)(?:º)?\s*andar", texto, re.IGNORECASE)
    dados['andar'] = andar_match.group(1) if andar_match else "0"

    # booleanos
    dados['piscina'] = "1" if re.search(r"piscinas?", texto, re.IGNORECASE) else "0"
    dados['varanda'] = "1" if re.search(r"varandas?", texto, re.IGNORECASE) else "0"
    dados['elevador'] = "1" if re.search(r"elevador", texto, re.IGNORECASE) else "0"

    return dados


def dividir_endereco(endereco_texto):
    if not endereco_texto or endereco_texto == "0":
        return "0", "0", "0", "0"

    m = re.match(r"^(.*?)\s*-\s*(.*?),\s*(.*?)\s*-\s*(.*)$", endereco_texto)
    if m:
        return m.group(1).strip(), m.group(2).strip(), m.group(3).strip(), m.group(4).strip()

    return endereco_texto, "0", "0", "0"


def get_address_from_soup(soup):
    tag = soup.select_one('p[data-testid="location-address"], div[data-testid="location-address"]')
    if tag:
        return tag.get_text(" ", strip=True)

    tag = soup.select_one('span[itemprop="streetAddress"]')
    if tag:
        return tag.get_text(" ", strip=True)

    return "0"


def extrair_informacoes(url):
    options = webdriver.ChromeOptions()
    # navegador visual
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")

    # remove logs chatos
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument("--log-level=3")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.maximize_window()
    driver.get(url)

    time.sleep(1)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

    html_completo = driver.page_source
    driver.quit()

    soup = BeautifulSoup(html_completo, "html.parser")
    texto = soup.get_text(" ", strip=True)

    dados = extrair_valores(texto)
    endereco = get_address_from_soup(soup)
    dados['endereco_completo'] = endereco

    rua, bairro, cidade, estado = dividir_endereco(endereco)
    dados['rua'] = rua
    dados['bairro'] = bairro
    dados['cidade'] = cidade
    dados['estado'] = estado

    return dados


def processar_links(arquivo_txt):
    if not os.path.isfile(arquivo_txt):
        print(f"{Fore.RED}Arquivo não encontrado: {arquivo_txt}")
        return []

    with open(arquivo_txt, 'r', encoding='utf-8') as f:
        links = [l.strip() for l in f if l.strip()]

    resultados = []
    for url in tqdm(links, desc=f"{Fore.CYAN}Processando links", unit="link"):
        try:
            dados = extrair_informacoes(url)
            dados['link'] = url
            resultados.append(dados)
        except Exception as e:
            print(f"{Fore.RED}Erro ao processar {url}: {e}")

    return resultados


def estilizar_excel(nome_arquivo):
    wb = load_workbook(nome_arquivo)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.freeze_panes = "A2"
    wb.save(nome_arquivo)


if __name__ == "__main__":
    limpar_console()
    print(Fore.GREEN + Style.BRIGHT + "=== COLETOR DE DADOS VIVAREAL ===\n")

    arquivo_default = r"C:\Users\kalebe\Desktop\vivareal.txt"
    caminho = input(f"{Fore.YELLOW}Informe o caminho do arquivo TXT (Enter para usar '{arquivo_default}'): ").strip()
    if not caminho:
        caminho = arquivo_default

    resultados = processar_links(caminho)

    if resultados:
        df = pd.DataFrame(resultados)
        colunas = ['preco', 'metragem', 'quartos', 'banheiros', 'vagas', 'suítes',
                   'andar', 'piscina', 'varanda', 'elevador',
                   'rua', 'bairro', 'cidade', 'estado', 'endereco_completo', 'link']
        df = df.reindex(columns=colunas)

        output_file = "resultados.xlsx"
        df.to_excel(output_file, index=False)
        estilizar_excel(output_file)

        print(Fore.GREEN + f"\nResultados salvos com sucesso em {output_file}.\n")
    else:
        print(Fore.RED + "Nenhum resultado para salvar.")
